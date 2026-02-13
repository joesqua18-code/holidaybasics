#!/usr/bin/env python3
import openpyxl
import csv

# Load category mapping
category_map = {}
with open('data/category_breakdown.csv', 'r') as f:
    reader = csv.DictReader(f)
    for row in reader:
        code = row['Category Code'].strip()
        category_map[code] = row['Description'].strip()

# Load Excel
wb = openpyxl.load_workbook('DEC ITEM MASTER.xlsx')
ws = wb.active

# Write CSV
with open('data/products.csv', 'w', newline='') as f:
    writer = csv.writer(f)
    writer.writerow(['STYLE', 'DESC', 'CATEGORY', 'LOT', 'QOH_CASES', 'UPC_SKU_2', 'PRICE_CS', 'PRICE_UNIT', 'SIZE'])

    for row in range(2, ws.max_row + 1):
        item_code = ws[f'A{row}'].value or ''
        desc = ws[f'B{row}'].value or ''
        uom = ws[f'C{row}'].value or ''
        prod_code = str(ws[f'D{row}'].value or '').strip()
        qty_raw = ws[f'L{row}'].value
        upc = str(ws[f'M{row}'].value or '').strip()

        # Map category
        category = category_map.get(prod_code)
        if not category and prod_code.endswith('-'):
            category = category_map.get(prod_code[:-1])
        if not category:
            category = 'MISC'

        # Clean quantity
        if isinstance(qty_raw, str):
            qty_raw = qty_raw.replace('-', '').strip()
        try:
            qty = int(float(qty_raw)) if qty_raw else 0
        except:
            qty = 0

        writer.writerow([
            item_code,
            desc,
            category,
            uom,
            qty,
            upc,
            '10.00',  # PRICE_CS placeholder
            '1.00',   # PRICE_UNIT placeholder
            ''        # SIZE empty
        ])

print(f"✓ Converted {ws.max_row - 1} products to data/products.csv")
